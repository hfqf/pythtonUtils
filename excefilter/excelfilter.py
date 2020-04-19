# !/usr/bin/python
# -*- coding: utf-8 -*-
# @Date : 2020/03/30
# @Author : hfqf123@126.com
# @Description:  提取PXXXXXXXXXXXXXX关键内容
import os
import sys
import xlrd
import xlsxwriter
import codecs
import re
import openpyxl
reload(sys)
sys.setdefaultencoding('utf-8')


excelPath = "11-K3.xls"

def start():
    if os.path.exists(excelPath):
        print(excelPath+"has existed!")
        data = xlrd.open_workbook('11-K3.xls')
        mywb = openpyxl.Workbook()
        sheet = mywb.active
        mywb.create_sheet('newSheet',0)
        sheet = mywb.get_sheet_by_name('newSheet')
    
        arr = getContacts(data)
        startFilter(arr,sheet)
        mywb.save('1.xls')
#        shutil.rmtree(mp3Path)
#        print("mp3Path has deleted!")
#        os.mkdir(mp3Path)
#        print("mp3Path has created once again!")
    else:
        print(excelPath+"还不存在，请检查!")
   

def getTable(data):
    table = data.sheet_by_index(1)
    return table

def getContacts(data):
    table = getTable(data)
    data = table.col(6, start_rowx=0, end_rowx=None)
    return data

def startFilter(arr,sheet):
    col = 1
    row = 1
    for item in arr:
        str = unicode(item).decode('unicode_escape').encode('utf-8')
        print re.sub(r'[^0-9A-Za-z\u4e00-\u9fa5]','',str)
        matchObj = re.search(r'P\d\d\d\d\d\d\d\d\d\d\d\d\d*',str, re.M|re.I)
        if matchObj:
           temp = matchObj.group()
           sheet.cell(column = col,row = row,value = temp)
        else:
           sheet.cell(column = col,row = row,value = 'No match')
        row = row+1
        
start()



